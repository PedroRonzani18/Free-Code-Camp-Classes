import os
import subprocess
from pathlib import Path

from openpyxl import load_workbook

wb = load_workbook('barchart.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

sheet['B8'] = '=SUM(B6:B7)'
sheet['B8'].style = 'Currency'

wb.save('barchart.xlsx')

application_path = str(Path(__file__).parent)
file_path = os.path.join(application_path, 'barchart.xlsx')
subprocess.run(['libreoffice', file_path])

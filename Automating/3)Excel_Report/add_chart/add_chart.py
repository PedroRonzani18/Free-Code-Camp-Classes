import os
import subprocess
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference


def create_excel():
    application_path = str(Path(__file__).parent)
    file_path = os.path.join(application_path, 'pivot_table.xlsx')
    wb = load_workbook(file_path)
    sheet = wb['Report']

    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    data = Reference(sheet, min_col=min_column + 1, min_row=min_row, max_row=max_row, max_col=max_column)
    categories = Reference(sheet, min_col=min_column, min_row=min_row + 1, max_row=max_row, max_col=max_column)

    barchart = BarChart()
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)

    sheet.add_chart(barchart, "B12")

    barchart.title = "Sales by product line"
    barchart.style = 3
    wb.save('barchart.xlsx')


def open_excel():
    application_path = str(Path(__file__).parent)
    file_path = os.path.join(application_path, 'barchart.xlsx')
    subprocess.run(['libreoffice', file_path])


if __name__ == "__main__":
    create_excel()
    open_excel()
